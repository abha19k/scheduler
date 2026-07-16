from pathlib import Path
import shutil
from typing import Dict, Any, List
import os

import pandas as pd
from openpyxl import load_workbook

from fastapi import FastAPI, UploadFile, File
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse
from pydantic import BaseModel, Field

from backend.optimizer_service import run_ga_scheduler
from backend.scenario_excel_store import (
    load_saved_scenarios_from_excel,
    save_scenario_definition_to_excel,
    load_scenario_definitions_from_excel,
    delete_scenario_from_excel,
)

from shutil import copyfile
from zipfile import BadZipFile



app = FastAPI(title="PlanWise Scheduler API")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

BASE_DIR = Path(__file__).resolve().parent.parent
UPLOAD_DIR = BASE_DIR / "backend" / "uploads"
OUTPUT_DIR = BASE_DIR / "backend" / "output"
STORAGE_DIR = BASE_DIR / "backend" / "storage"

UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
STORAGE_DIR.mkdir(parents=True, exist_ok=True)

IMPORTED_WORKBOOK_PATH = STORAGE_DIR / "imported_workbook.xlsx"


class DowntimeRequest(BaseModel):
    MachineID: str
    StartTime: str
    EndTime: str
    Reason: str = "Downtime"


class RunScenarioRequest(BaseModel):
    scenarioId: str = "BASE"
    parameterOverrides: Dict[str, Any] = Field(default_factory=dict)
    downtimes: List[DowntimeRequest] = Field(default_factory=list)
    calendarOverrides: Dict[str, Any] = Field(default_factory=dict)
    machineOverrides: Dict[str, Any] = Field(default_factory=dict)
    objectiveOverrides: Dict[str, Any] = Field(default_factory=dict)


class SaveManualChangesRequest(BaseModel):
    scenarioId: str
    plannedTasks: List[Any]
    manualChanges: List[Any]


class SaveScenarioDefinitionRequest(BaseModel):
    ScenarioID: str
    ScenarioName: str
    BaseScenarioID: str | None = None
    Description: str = ""
    ParameterOverrides: Dict[str, Any] = Field(default_factory=dict)
    CalendarOverrides: Dict[str, Any] = Field(default_factory=dict)
    MachineOverrides: Dict[str, Any] = Field(default_factory=dict)
    ObjectiveOverrides: Dict[str, Any] = Field(default_factory=dict)
    Downtimes: List[Any] = Field(default_factory=list)


@app.get("/")
def health():
    return {"status": "PlanWise Scheduler Backend Running"}


@app.post("/run-ga")
def run_ga(request: RunScenarioRequest):
    return run_ga_scheduler(
    scenario_id=request.scenarioId,
    excel_file=str(IMPORTED_WORKBOOK_PATH),
    custom_parameter_overrides=request.parameterOverrides,
    custom_downtimes=[d.model_dump() for d in request.downtimes],
    custom_calendar_overrides=request.calendarOverrides,
    custom_machine_overrides=request.machineOverrides,
    custom_objective_overrides=request.objectiveOverrides,
    )



@app.post("/api/upload-excel")
async def upload_excel(file: UploadFile = File(...)):
    file_path = IMPORTED_WORKBOOK_PATH

    with open(file_path, "wb") as buffer:
        shutil.copyfileobj(file.file, buffer)

    excel = pd.ExcelFile(file_path)
    preview = {}

    for sheet_name in excel.sheet_names:
        df = pd.read_excel(file_path, sheet_name=sheet_name)

        preview[sheet_name] = {
            "columns": list(df.columns),
            "rows": df.fillna("").to_dict(orient="records"),
            "previewRows": df.head(20).fillna("").to_dict(orient="records"),
            "totalRows": len(df),
        }

    return {
        "message": "Excel uploaded successfully",
        "filePath": str(file_path),
        "sheets": preview,
    }

@app.get("/api/imported-data")
def get_imported_data():
    file_path = IMPORTED_WORKBOOK_PATH

    if not file_path.exists():
        return {
            "message": "No imported Excel found",
            "sheets": {}
        }

    excel = pd.ExcelFile(file_path)
    preview = {}

    for sheet_name in excel.sheet_names:
        df = pd.read_excel(file_path, sheet_name=sheet_name)

        preview[sheet_name] = {
            "columns": list(df.columns),
            "rows": df.fillna("").to_dict(orient="records"),
            "previewRows": df.head(20).fillna("").to_dict(orient="records"),
            "totalRows": len(df),
        }

    return {
        "message": "Imported Excel loaded successfully",
        "filePath": str(file_path),
        "sheets": preview,
    }


@app.post("/api/optimize")
async def optimize(
    file: UploadFile = File(...),
    scenarioId: str = "BASE"
):
    file_path = UPLOAD_DIR / file.filename

    with open(file_path, "wb") as buffer:
        shutil.copyfileobj(file.file, buffer)

    return run_ga_scheduler(
        scenario_id=scenarioId,
        excel_file=str(file_path),
    )


@app.get("/api/scenarios")
def get_saved_scenarios():
    return {
        "scenarios": load_saved_scenarios_from_excel()
    }


@app.get("/api/scenario-definitions")
def get_scenario_definitions():
    return {
        "definitions": load_scenario_definitions_from_excel()
    }


@app.post("/api/save-scenario-definition")
def save_scenario_definition(request: SaveScenarioDefinitionRequest):
    save_scenario_definition_to_excel(
        request.model_dump()
    )

    return {
        "success": True
    }

@app.delete("/api/scenario/{scenario_id}")
def delete_scenario(scenario_id: str):
    if scenario_id == "BASE":
        return {
            "success": False,
            "message": "Base Scenario cannot be deleted."
        }

    delete_scenario_from_excel(scenario_id)

    return {
        "success": True,
        "deletedScenarioId": scenario_id
    }


@app.post("/api/save-manual-changes")
def save_manual_changes(request: SaveManualChangesRequest):
    excel_path = STORAGE_DIR / "planwise_scenario_store.xlsx"
    temp_path = STORAGE_DIR / "planwise_scenario_store.tmp.xlsx"

    if not excel_path.exists():
        raise ValueError("Scenario Excel store not found.")

    planned_df = pd.DataFrame(request.plannedTasks)
    manual_df = pd.DataFrame(request.manualChanges)

    planned_sheet_name = f"{request.scenarioId}_planned"[:31]
    manual_sheet_name = f"{request.scenarioId}_manual"[:31]

    try:
        copyfile(excel_path, temp_path)

        workbook = load_workbook(temp_path)

        for sheet_name in [planned_sheet_name, manual_sheet_name]:
            if sheet_name in workbook.sheetnames:
                del workbook[sheet_name]

        workbook.save(temp_path)
        workbook.close()

        with pd.ExcelWriter(
            temp_path,
            engine="openpyxl",
            mode="a",
            if_sheet_exists="replace"
        ) as writer:
            planned_df.to_excel(
                writer,
                sheet_name=planned_sheet_name,
                index=False
            )

            manual_df.to_excel(
                writer,
                sheet_name=manual_sheet_name,
                index=False
            )

        temp_path.replace(excel_path)

        return {"success": True}

    except BadZipFile:
        raise ValueError(
            "Scenario Excel store is corrupted. Delete planwise_scenario_store.xlsx and rerun the optimizer to recreate it."
        )

    finally:
        if temp_path.exists():
            temp_path.unlink(missing_ok=True)

@app.get("/gantt-image")
def gantt_image():
    possible_files = sorted(
        OUTPUT_DIR.glob("*_ga_aps_gantt.png"),
        key=lambda p: p.stat().st_mtime,
        reverse=True
    )

    gantt_path = (
        possible_files[0]
        if possible_files
        else OUTPUT_DIR / "ga_aps_gantt.png"
    )

    return FileResponse(
        gantt_path,
        media_type="image/png"
    )


@app.get("/download-excel")
def download_excel():
    possible_files = sorted(
        OUTPUT_DIR.glob("*_ga_aps_solution.xlsx"),
        key=lambda p: p.stat().st_mtime,
        reverse=True
    )

    excel_path = (
        possible_files[0]
        if possible_files
        else OUTPUT_DIR / "ga_aps_solution.xlsx"
    )

    return FileResponse(
        excel_path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=excel_path.name,
    )